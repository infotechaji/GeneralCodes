"""
File version : v1.0
Document reference version : v1.0

History : 
        v1.0    - 12/12/2020  - intial version 

"""

import pytesseract,sys,os,re
from PIL import Image
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from CONFIG import *
# pytesseract.pytesseract.tesseract_cmd = 'C:\Program Files\Tesseract-OCR/tesseract.exe'
pytesseract.pytesseract.tesseract_cmd = TESSERACT_PATH
HEADERS=['FILE_NAME','TRAN_ID','TRAN_DATE','TRAN_TIME','TRAN_STATUS','BILLER_NAME','BILLER_ID','K_NUMBER','PAYMENT_MODE','MOBILE','AMOUNT','BILL_DATE','CONV_FEES','TOTAL_AMOUNT']

def apply_to_list(input_list,make_upper = False, make_lower= False,make_int= False,make_string=False):
    # added to make the list comprehension process easy  , added on 15/10/2020
    temp_list = input_list
    if make_string: return [str(i) for i in temp_list]
    elif make_upper: return [str(i).upper() for i in temp_list]
    elif make_lower: return [str(i).lower() for i in temp_list]
    elif make_int: return [int(i) for i in temp_list]
    return temp_list

def handle_extension(input_sp, extension_to_add='.sql',developer_mode = False):
    status = False
    DIRECTORY, extracted_file_name = os.path.split(input_sp)
    sp_name = None
    file_name = None
    try:
        sp_name = str(extracted_file_name).split('.')[0]
        file_name = str(sp_name) + str(extension_to_add)
        status = True
    except Exception as e:
        developer_print('Error in handling name :',e)
        status = False
    return {
            'no_extension': sp_name
            , 'new_extension': file_name
            , 'directory': DIRECTORY
            , 'file_name': extracted_file_name
            }


def write_excel(work_book_name, rows=[], headers=''):
    try:
        try:
            book = openpyxl.load_workbook(work_book_name)
            sheet = book.active

        # print('Opening excisting file !')
        except Exception as e:
            # book = Workbook(write_only=True)
            # print('Error , Creating new excel file ')
            book = Workbook(write_only=True)
            sheet = book.create_sheet()
            if headers:
                sheet.append(headers)

        for row in rows:
            sheet.append(row)

        book.save(work_book_name)
    except Exception as e:
        print ('Please close the output excel file  !!!! and rerun the application')
        exit()
    return True


def process_image_to_text(file_lines, file_name='', developer_mode=False):
    # headers = ['FILE_NAME', 'TRAN_ID', 'TRAN_DATE','TRAN_TIME', 'TRAN_STATUS', 'BILLER_NAME', 'BILLER_ID', 'K_NUMBER',
    #            'PAYMENT_MODE', 'AMOUNT', 'BILL_DATE', 'CONV_FEES', 'TOTAL_AMOUNT']

    tran_id = ''
    tran_date = ''
    tran_time = ''
    tran_status = ''


    biller_name = ''
    biller_id = ''
    k_number = ''
    payment_mode = ''
    mobile = ''
    amount = ''
    bill_date = ''
    conv_fees = ''
    total_amount = ''

    for index, looping_line in enumerate(file_lines):
        each_line = looping_line
        if developer_mode:
            # print(index, each_line)
            # each_line
            pass

        # t_date_pattern = 'Txn. Date/Time (\w.+)'
        
        if not k_number:
            try:
                pattern = '\s(\d{12})\s'
                total_ids = re.findall(pattern, str(each_line), re.IGNORECASE)
                if total_ids:
                    if developer_mode:
                        print('k-number total_ids :', total_ids)
                    k_number = total_ids[0]
                if developer_mode:
                    print('extracted k-number :', k_number)
            except Exception as e:
                print ('Error while getting k_number :',e)

        if k_number:
            each_line = each_line.replace(k_number,'').strip()
        if not amount:
            try:
                pattern = '(\d\.00)'
                pattern2 = '\d.+\.00'
                amounts_list = []
                for sub_line in each_line.split('\n'):
                    total_ids = re.findall(pattern, str(sub_line), re.IGNORECASE)
                    if total_ids:
                        if developer_mode:
                            print('Amount total_ids :', total_ids)
                        splits = sub_line.split()
                        for i in splits:
                            if '.00' in i.lower():
                                amount_suspect = i
                                if developer_mode:
                                    print('amount_suspect:', amount_suspect)
                                second_extract = re.findall(pattern2, str(amount_suspect), re.IGNORECASE)
                                if second_extract:
                                    if developer_mode:
                                        print('second_extract:', second_extract)
                                    amount = second_extract[0]
                                    amounts_list.append(amount)
                if developer_mode:
                    print('amounts_list :',amounts_list)
                if len(amounts_list)==2:
                    amount = amounts_list[0]
                    total_amount = amounts_list[1]
            except: pass
        each_line = each_line.replace('\r',' ')
        each_line = each_line.replace('\n',' ')
        # print ('each_line before processing :',each_line)
        if not tran_id:
            try:
                tran_id_first =''
                tran_id_second =''
                pattern_a_12 = '\s(\w{12})\s' # 8
                pattern_a_14 = '\s(\w{14})\s' # 6
                total_ids = []
                length_of_second = 8
                if developer_mode:
                    print('each_line :',each_line)
                total_ids_12 = re.findall(pattern_a_12, str(each_line), re.IGNORECASE)
                total_ids_14 = re.findall(pattern_a_14, str(each_line), re.IGNORECASE)
                if developer_mode:
                    print('total_ids_12 :',total_ids_12)
                    print('total_ids_14 :',total_ids_14)
                if total_ids_12 and total_ids_12[0].isnumeric()==False and total_ids_12[0].isalpha()==False:
                        if developer_mode:
                            print('Detected type : 12')
                        length_of_second = 8
                        total_ids = total_ids_12
                elif total_ids_14 and total_ids_14[0].isnumeric()==False and total_ids_14[0].isalpha()==False:
                    if developer_mode:
                        print('Detected type : 14')
                    length_of_second = 6
                    total_ids = total_ids_14
                else:
                    if developer_mode:
                        print('tran ids are not detected')

                if total_ids:
                    if developer_mode:
                        print('tran_id(1) total_ids :', total_ids)
                    tran_id_first = total_ids[0]
                    if developer_mode:
                        print('extracted tran_id-1 :', tran_id_first)
                # else:
                #     pattern = '\s(\w{14})\s'
                #     total_ids = re.findall(pattern, str(each_line), re.IGNORECASE)
                #     if total_ids:
                        
                #         if developer_mode:
                #             print('tran_id(1) total_ids :', total_ids)
                #         tran_id_first = total_ids[0]
                #         if developer_mode:
                #             print('extracted tran_id-1 :', tran_id_first)


                pattern2 = '\s(\d{'+str(length_of_second)+'})\s'
                total_ids = re.findall(pattern2, str(each_line), re.IGNORECASE)
                if total_ids:
                    if developer_mode:
                        print('tran_id total_ids :', total_ids)
                    if total_ids[0].isnumeric():
                        tran_id_second = total_ids[0]
                if developer_mode:
                    print('extracted tran_id-2 :', tran_id_first)
                tran_id = tran_id_first+' '+tran_id_second

                if developer_mode:
                    print('tran_id :', tran_id)
            except Exception as e:
                if developer_mode:
                    print('Error while extracting tran_status :',e)
            

        try:
            if tran_id:
                each_line = each_line.replace(tran_id.split()[0],'').replace(tran_id.split()[1],'')
        except Exception as e:
                if developer_mode:
                    print('Error while deleting the tran id details :',e)

        if not biller_id:
            try:
                pattern = '\s(\w{14})\s'
                total_ids = re.findall(pattern, str(each_line), re.IGNORECASE)
                if total_ids:
                    if developer_mode:
                        print('biller_id total_ids :', total_ids)
                    biller_id = total_ids[0]
                if developer_mode:
                    print('extracted biller_id :', biller_id)
            except Exception as e:
                if developer_mode:
                    print('Error while extracting biller_id :',e)
        
            # Payment Mode (\w+) payment mode
        if not payment_mode:
            try:
                acceptable_payments =['credit']
                pattern = 'Payment Mode (\w+)'
                total_ids = re.findall(pattern, str(each_line), re.IGNORECASE)
                if total_ids:
                    if developer_mode:
                        print('payment_mode total_ids :', total_ids)
                    if total_ids[0].lower() in acceptable_payments:
                        payment_mode = total_ids[0]
                    if developer_mode:
                        print('extracted payment_mode :', payment_mode)
                    if not payment_mode:
                       for i in  acceptable_payments:
                        if re.findall(i, str(each_line), re.IGNORECASE):
                            payment_mode = i.capitalize()
                            break
            except Exception as e:
                if developer_mode:
                    print('Error while extracting payment_mode :',e)
        if not bill_date:
            try:
                pattern = '\d{2}/\d{2}/\d{4}'
                total_ids = re.findall(pattern, str(each_line), re.IGNORECASE)
                if total_ids:
                    if developer_mode:
                        print('bill date total_ids:', total_ids)
                    bill_date = str(total_ids[0])
                if developer_mode:
                    print('extracted bill_date :', bill_date)
            except Exception as e:
                if developer_mode:
                    print('Error while extracting bill_date :',e)
        if not tran_date:
            try:
                pattern = '\w+\s+\d{1,2},\d{4}'
                pattern_time = '(\d{1,2}:\d{1,2})\s+(\w{2})\s'
                total_ids = re.findall(pattern, str(each_line), re.IGNORECASE)
                total_ids_time = re.findall(pattern_time, str(each_line), re.IGNORECASE)
                if total_ids:
                    if developer_mode:
                        print('tran_date total_ids :', total_ids)
                    tran_date = str(total_ids[0])
                if developer_mode:
                    print('extracted tran_date :', tran_date)
                if total_ids_time:
                    if developer_mode:
                        print('tran_time total_ids :', total_ids_time)
                    tran_time = str(total_ids_time[0][0]+total_ids_time[0][1])
                if developer_mode:
                    print('extracted tran_time :', tran_time)
            except Exception as e:
                if developer_mode:
                    print('Error while extracting bill_date :',e)
            
        if not tran_status:
            try:
                accepted_status = ['success','failed']
                pattern = '\s([A-Z]{6,7})\s'
                total_ids = re.findall(pattern, str(each_line))
                # total_ids = re.findall(pattern, str(each_line), re.IGNORECASE)
                if total_ids:
                    if developer_mode:
                        print('Status total_ids :', total_ids)
                    for each_res in total_ids:
                        if each_res.lower() in accepted_status:
                            tran_status = each_res
                if developer_mode:
                    print('extracted tran_status :', tran_status)
            except Exception as e:
                if developer_mode:
                    print('Error while extracting tran_status :',e)
        
        
        if not biller_name:
            SUSPECT_JUNKS = ['Biller Name','Biller','Name','ID','K Number','Number','Amount','Bill Date','Bill','Date','Convenience Fees','Total Amount','Your transaction is completed successfully','BIBRABAT','INVOICE']
            try:
                hint = ''
                hint_first_letter = ''

                pattern = '\((\w+)\)'
                total_ids = re.findall(pattern, str(each_line), re.IGNORECASE)
                if total_ids:
                    hint = str(total_ids[0])
                    hint_first_letter = hint[0]
                    try:
                        suspect = ' '.join(each_line.split(hint)[0].split()[-8:])+str(hint)+')'
                    except :
                        print(len(each_line.split(hint)[0].split()))
                        suspect = ' '.join(each_line.split(hint)[0].split()[-8:])+str(hint)+')'
                    # print ('suspect string:',suspect)
                    if developer_mode:
                        print('biller_name hint :', hint)
                        print('biller_name hint_first_letter :', hint_first_letter)
                if hint:
                    for each_junk in SUSPECT_JUNKS:
                        suspect = suspect.replace(each_junk,'').strip()
                    pattern2 = '('+str(hint_first_letter)+'\w.+\('+str(hint)+'\))'
                    if developer_mode:
                            print('biller_name developed pattern :', pattern2)
                    total_ids2 = re.findall(pattern2, str(suspect), re.IGNORECASE)
                    if developer_mode:
                        print ('total_ids2 - findall :',total_ids2)
                    if total_ids2:
                        if developer_mode:
                            print('biller_name total_ids :', total_ids2)
                        biller_name = total_ids2[0]
                    if developer_mode:
                        print('extracted biller_name :', biller_name)
            except Exception as e:
                if developer_mode:
                    print('Error while extracting biller_name :',e)
        if not mobile:
            try:
                pattern = '\s(\d{10})\s'
                total_ids = re.findall(pattern, str(each_line), re.IGNORECASE)
                if total_ids:
                    if developer_mode:
                        print('mobile-number total_ids :', total_ids)
                    mobile = total_ids[0]
                if developer_mode:
                    print('extracted mobile-number :', mobile)
            except Exception as e:
                if developer_mode:
                    print('Error while extracting bill_date :',e)
        
        # Exception cases handling 
        
        # if not biller_name:
        #     pattern = '(\w.+)\s(\w+)\s(\d{12})'
        #     total_ids = re.findall(pattern, str(each_line), re.IGNORECASE)
        #     if total_ids:
        #         if developer_mode:
        #             print('Biller id  total_ids :', total_ids)
        #         biller_name2 = str(total_ids[0][0]).strip('\n')
        #         biller_id = str(total_ids[0][1]).strip()
        #         k_number = str(total_ids[0][2]).strip()
        #         payment_mode = str(file_lines[index + 1]).strip()
        #         biller_name = str(file_lines[index - 1]).strip() + str(biller_name2)
        #     if developer_mode:
        #         print('extracted tran_status :', tran_status)
        #         print('extracted biller_name :', biller_name)
        #         print('extracted biller_id :', biller_id)
        #         print('extracted k_number :', k_number)
        #         print('extracted payment_mode :', payment_mode)
        
        
    # if  'txn. id' in each_line.lower():
    # 	t_id = each_line.replace('Txn. ID ','').strip()

    # elif  'status' in each_line.lower():
    # 	t_id = each_line.replace('Status ','').strip()

    res = {
        'tran_id': tran_id
        , 'tran_date': tran_date
        , 'tran_time': tran_time
        , 'tran_status': tran_status

        , 'biller_name': biller_name
        , 'biller_id': biller_id
        , 'k_number': k_number
        , 'payment_mode': payment_mode
        , 'mobile': mobile

        , 'amount': amount
        , 'bill_date': bill_date
        , 'conv_fees': conv_fees
        , 'total_amount': total_amount
        , 'file_name': file_name
    }

    # # print ('res :',res)
    # full_log = [res['file_name'],res['tran_id'], res['tran_date']#, res['tran_time']
    # 			,res['tran_status'],
    # 			res['biller_name'], res['biller_id'],res['k_number'],res['payment_mode'], res['amount'] ,
    # 			res['bill_date'], res['conv_fees'],res['total_amount']
    # 			]
    # full_log_text = '\t'.join(apply_to_list(full_log, make_string=True)) + '\n'
    # write_into_file(file_name='Google_pay_log.txt', contents=add_timestamp(full_log_text), mode='a')
    # result_list.append(res)

    full_log = [res['file_name'], res['tran_id'], res['tran_date'],res['tran_time'],
                res['tran_status'],
                res['biller_name'], res['biller_id'], res['k_number'], res['payment_mode'],res['mobile'], res['amount'],
                res['bill_date'], res['conv_fees'], res['total_amount']
                ]
    full_log_text = apply_to_list(full_log, make_string=True)
    # if developer_mode:

    dt_string = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    # print("date and time =", dt_string)
    full_log_text_log = full_log_text.copy()
    full_log_text_log.insert(0, str(dt_string))
    headers_log = HEADERS.copy()
    headers_log.insert(0, 'TIMESTAMP')
    write_excel(work_book_name='TOTAL_FILES_LOG.xlsx', rows=[full_log_text_log], headers=headers_log)
    write_excel(work_book_name='Extracted_text.xlsx', rows=[full_log_text], headers=HEADERS)
    # with open('results_dev_mode.txt', mode='a') as file:
    #     file.write('\t'.join(full_log_text_log)+'\n')
    return res

def get_text_from_image(input_image,developer_mode = True,save_files = False,text_directory = 'text_files',dir = ''):
    text_file_name = handle_extension(input_image, extension_to_add='.txt')['new_extension']
    text_file_path =os.path.join(dir,text_directory)
    if not os.path.exists(text_file_path):os.makedirs(text_file_path)
    full_text_file_path = os.path.join(text_file_path,text_file_name)
    if developer_mode:
        print('full_text_file_path:',full_text_file_path)


    if os.path.exists(full_text_file_path):
        result = str(open(full_text_file_path).read())
    else:
        img = Image.open(os.path.join(dir,input_image))
        result = pytesseract.image_to_string(img)
        if save_files:
            with open(full_text_file_path, mode='w') as file:
                file.write(result)
            if developer_mode:
                print('Text file saved  :', full_text_file_path)
    # print(type(result))
    # print(result)
    # file_lines = list(filter(None,result.split('\n')))
    # print('Files lines :',file_lines)
    # try:
    if True:
        process_image_to_text([result],file_name= input_image,developer_mode= developer_mode)
        print(' Text extracted!')
    # except Exception as e:
    #     print(' Text not extracted! ( Rerun the application )')

    return True



if __name__ =="__main__":
    # with open('results_dev_mode.txt', mode='w') as file:
    #     file.write('\t'.join(HEADERS)+'\n')

    # input_directory = sys.argv[1]
    # save_file_option = 'yes'
    while True:
        input_directory = input('Input directory that contains images :\t')
        save_file_option = input('Do you want save converted text files? Yes/No :\t')
        
        input_directory = input_directory.strip()
        if not os.path.exists(input_directory):
            print('Please provide a valid directory')
            exit()
        try:
            os.remove('Extracted_text.xlsx')
        except:pass
        file_count =0 
        save_file = False
        if save_file_option.strip() =='yes':
            save_file =True
        for root, dirs, files in os.walk(input_directory):
            input_files = []
            for file_name in files:
                file_count+=1
                print ('Processing ... {0}/{1}, {2} :'.format(file_count,len(files),file_name),end ='')
                get_text_from_image(dir=input_directory, input_image = file_name,developer_mode=False,save_files = save_file)
                # input('Proceed further ?')
            break
        if 'exit' in str(input('Press "exit" to close ')).lower():break
