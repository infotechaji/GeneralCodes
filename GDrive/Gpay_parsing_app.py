"""
Definition : Performs Upload/Download files to the googele drive 
Version : v1.1.2
History : 
		  v1.1   - 05/07/2018 - Initial version
		  v1.1.1 - 05/08/2018 - Function download_files is updated to download files to a specific directory 
		  v1.1.2 - 05/08/2018 - Functions search_file,delete_file are added to optime the search 
		  v2.0   - 11/15/2018 - P tag with Page number is added if it present only.
Pending:

Issues :

"""


#from __future__ import print_function # comment this line to use python 2
from googleapiclient import discovery
from googleapiclient import errors
from googleapiclient import http
# from apiclient import *

from googleapiclient.http import MediaFileUpload
from googleapiclient.http import MediaIoBaseDownload
from httplib2 import Http
from oauth2client import file, client, tools
from get_html_file import *
import os,sys
import io,re
from datetime import datetime

# sys.path.insert(1, 'G:\\Ajith\\OtherFiles\\common')

# from Help_print import *
# from CustomisedFileOperation import *
# from  import *

import openpyxl
from openpyxl import Workbook


def write_excel(work_book_name, rows=[],headers =''):
	
	try:
		book = openpyxl.load_workbook(work_book_name)
		sheet = book.active
		
		# print('Opening excisting file !')
	except Exception as e :
		# book = Workbook(write_only=True)
		# print('Error , Creating new excel file ')
		book = Workbook(write_only = True)
		sheet = book.create_sheet()
		if headers:
			sheet.append(headers)
	

	for row in rows:
		sheet.append(row)
	
	book.save(work_book_name)
	return True

FOLDER_MIME = 'application/vnd.google-apps.folder'
class GoogleDrive():
	def __init__(self,client_key_json,developer_mode=True,out_file = 'Extracted_text.xlsx'):
		self.developer_mode=developer_mode
		self.SCOPES = 'https://www.googleapis.com/auth/drive'
		self.store = file.Storage('storage.json')
		self.creds = self.store.get()
		self.mimeList=[{'format':'HTML','mimeType':'text/html'},{'format':'zipped','mimeType':'application/zip'},{'format':'Plain text','mimeType':'text/plain'},{'format':'Rich text','mimeType':'application/rtf'},{'format':'Open Office doc','mimeType':'application/vnd.oasis.opendocument.text'},{'format':'PDF','mimeType':'application/pdf'},{'format':'MS Word document','mimeType':'application/vnd.openxmlformats-officedocument.wordprocessingml.document'},{'format':'EPUB','mimeType':'application/epub+zip'},{'format':'MS Excel','mimeType':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'},{'format':'Open Office sheet','mimeType':'application/x-vnd.oasis.opendocument.spreadsheet'},{'format':'PDF','mimeType':'application/pdf'},{'format':'CSV','mimeType':'text/csv'},{'format':'TSV','mimeType':'text/tab-separated-values'},{'format':'JPEG','mimeType':'image/jpeg'},{'format':'PNG','mimeType':'image/png'},{'format':'SVG','mimeType':'image/svg+xml'},{'format':'PDF','mimeType':'application/pdf'},{'format':'MS PowerPoint','mimeType':'application/vnd.openxmlformats-officedocument.presentationml.presentation'},{'format':'Open Office presentation','mimeType':'application/vnd.oasis.opendocument.presentation'},{'format':'PDF','mimeType':'application/pdf'},{'format':'Plain text','mimeType':'text/plain'},{'format':'JSON','mimeType':'application/vnd.google-apps.script+json'}]
		if not self.creds or self.creds.invalid:
			self.flow = client.flow_from_clientsecrets(client_key_json, self.SCOPES)
			self.creds = tools.run_flow(self.flow, self.store)
		self.DRIVE = discovery.build('drive', 'v2', http=self.creds.authorize(Http()))
		try:
			os.remove(out_file)
		except Exception as e:
			print('Error while removing file :',out_file)

	
	def upload_files(self,file_name='',directory='',folder_name='',convert=False): # uploads the file to the drive , If folder name is given then uploads file to the folder.
		print_prefix='upload_files\t'
		if not file_name and not directory:
			print ('Either File_name or Directory Should be specified !!')
			exit()
		if file_name:
			metadata = {'title': file_name.split(os.sep)[-1]}
			if folder_name:
				item=self.create_folder(folder_name=folder_name)
				folder_id=item['id']
				#print 'folder created :',folder_id
				metadata['parents']=[{'id': folder_id}]
			print('metadata :',metadata)
			try:
				res = self.DRIVE.files().insert(convert=convert,body=metadata,media_body=file_name).execute()
				# print(('File ',file_name ,(res['id']),'Uploaded Successfully!!!'))
				return res
			except errors.HttpError as error:
				print('An error occurred: %s' % error)
				return None
		if directory:
			for root, dirs, files in os.walk(directory):
				file_count=0
				if folder_name:
					item=self.create_folder(folder_name=folder_name)
					folder_id=item['id']
					# print('Folder ID :',folder_id)
					#metadata['parents']=[{'id': folder_id}]
				for file_name in files:
					each_file=os.path.join(directory,file_name)
					file_count+=1
					metadata = {'title': file_name.split(os.sep)[-1]}
					if folder_name and folder_id:
						metadata['parents']=[{'id': folder_id}]
					res = self.DRIVE.files().insert(convert=convert, body=metadata,media_body=each_file).execute()
					print(('File ',file_name ,'Uploaded Successfully!!!'))
		return True
	def upload_images(self,file_name='',directory='',MIMETYPE='text/plain',convert=True,extension='.txt',folder_name='image-conversion',output_directory='',developer_mode= False):
		output_directory=os.path.join(directory,'text_files')
		success = []
		errored = []
		if not os.path.exists(output_directory):os.makedirs(output_directory)
		
		if folder_name:
			item=self.create_folder(folder_name=folder_name)
			folder_id=item['id']
		
		for root, dirs, files in os.walk(directory):
			file_count=0
			for file_name in files:
				file_count+=1
				print ('Processing files:{0}/{1}, {2} '.format(file_count,len(files),file_name),end ="")
				each_file=os.path.join(directory,file_name)
				
				text_file_name_tmp=file_name.split('.')[0]
				text_file_name = text_file_name_tmp+'.txt'
				# print ('Looking for the file name :',text_file_name)
				if not os.path.exists(os.path.join(output_directory,text_file_name)):
					metadata = {'title': file_name.split(os.sep)[-1]}
					if folder_name and folder_id:
						metadata['parents']=[{'id': folder_id}]
					#print 'each_file:',each_file
					#print 'convert:',convert
					try:
						res = self.DRIVE.files().insert(convert=convert, body=metadata,media_body=each_file).execute()
						# print(file_count,'Uploaded !!:',file_name.split(os.sep)[-1])
						exp_results=gd.export_file(file_id=res['id'],output_mimeType=MIMETYPE,output_directory=output_directory)
						# if exp_results:
						# 	file_path=os.path.join(output_directory,exp_results['local_file_name'])
						# 	print('file_path :',file_path)
						
						success.append(file_name)
						
				
						# ht=HtmlActivites(file_path)
						# ht.write_headers()
						print (' Status: Success! ',end ='')
					except:
						errored.append(file_name)
						print (' Status: Errored! ',end ='')

				else:
					print (' Status: Skipped! ',end ='')

				full_results = google_pay_format_data(input_directory = output_directory,input_files=[text_file_name],developer_mode = developer_mode)['results']

			break
		return {'success':success
				,'errored':errored
				,'data_extracted':errored
				}

	def upload_and_export_files(self,file_name='',directory='',MIMETYPE='text/plain',convert=True,extension='.txt',folder_name='',output_directory='',developer_mode= True):
		success =[]
		errored =[]
		if file_name:
			res=gd.upload_files(file_name=file_name,convert=convert)
			if res:
				print(gd.export_file(file_id=res['id'],output_mimeType=MIMETYPE,output_directory=output_directory))
		elif directory:
			output_directory=os.path.join(directory,'text_files')
			try: 
				os.makedirs(output_directory)
			except OSError:
				#if not os.path.isdir(path):
				pass	
			
			for root, dirs, files in os.walk(directory):
				file_count=0
				if folder_name:
					item=self.create_folder(folder_name=folder_name)
					folder_id=item['id']
					if developer_mode:
						# print('Folder ID :',folder_id)
						pass
					#metadata['parents']=[{'id': folder_id}]
				
				for file_name in files:
					file_count+=1
					print ('Processing files :{0}/{1},  '.format(file_count,len(files)))
					each_file=os.path.join(directory,file_name)
					
					metadata = {'title': file_name.split(os.sep)[-1]}
					if folder_name and folder_id:
						metadata['parents']=[{'id': folder_id}]
					#print 'each_file:',each_file
					#print 'convert:',convert
					try:
						res = self.DRIVE.files().insert(convert=convert, body=metadata,media_body=each_file).execute()
						print(file_count,'Uploaded !!:',res['title'])
						exp_results=gd.export_file(file_id=res['id'],output_mimeType=MIMETYPE,output_directory=output_directory)
						if exp_results:
							file_path=os.path.join(output_directory,exp_results['local_file_name'])
							print('file_path :',file_path)
						success.append(file_name)
						# ht=HtmlActivites(file_path)
						# ht.write_headers()
					except:
						errored.append(file_name)
				break
		return {'success':success
				,'errored':errored}

	def export_file(self,output_mimeType='application/pdf',file_name='',file_id='',output_directory=os.getcwd()):
		if file_name and not file_id:
			items=self.search_file(search_key=file_name)
			extension='.txt'
			if len(items)==1:
				item=items[0]
				item_count=0
				# for i in item :
				# 	item_count+=1
				# 	print item_count,i,item[i]
				file_id=item.get('id')
				o_file_name=item['title'].split('.')[0]
		if file_id and not file_name:
			extension='.txt'
			result_dict=self.DRIVE.files().get(fileId=file_id).execute()
			# for i in result_dict:
			# 	print i ,result_dict[i]
			# export_links=result_dict['exportLinks']
			o_file_name=result_dict['title']
			file_name=o_file_name.split('.')[0]
			data = self.DRIVE.files().export(fileId=file_id, mimeType=output_mimeType).execute()
			#print 'data',data
			output_file_name = file_name+extension
			outfile=os.path.join(output_directory,output_file_name)
			#print 'outfile:',outfile
			if os.path.isfile(outfile):
				print("ERROR, %s already exist" % outfile)
			else:
				with open(outfile, 'wb') as f:
					f.write(data)
				# print("OK")
				return {'local_file_name':outfile}


	# def get_file_details(self,search_key):
	# 	return self.retrieve_all_files(search_key=search_key)
	def download_files(self,file_name='',file_id='',output_directory=os.getcwd(),is_folder=False,mimeType=''): # pending 
		result_dict={}
		if file_id:
			result_dict = self.DRIVE.files().get(fileId=file_id).execute()
		elif file_name:
			print('Given File name:',file_name)
			items=self.search_file(search_key='file_name',mimeType=mimeType)
			print('len(items):',len(items))
			if len(items)==1:
				result_dict=items[0]
			elif len(items)>1:
				print('No of Matches found for \"'+file_name+'\" Count :'+str(len_items))
				for index,item in enumerate(items):
					print('Found file: %d : %s (%s) - %s' % (index,item.get('title'), item.get('id'),item.get('mimeType')))
				option=input('To Download all files Press \"Yes\" or Press the index number to Download : ')
				if option.lower()=='yes': # To download all selected Files!!!
					for result_dict in items:
						output_file_name=result_dict['originalFilename']
						outfile=os.path.join(output_directory,output_file_name)
						print("downloading %s" % result_dict.get('title'))
						resp, content = self.DRIVE._http.request(download_url)
						if resp.status == 200:
							if os.path.isfile(outfile):
								print("ERROR, %s already exist" % outfile)
							else:
								with open(outfile, 'wb') as f:
									f.write(content)
								print("OK")
				elif option.isdigit():
					result_dict=items[int(option)]
					output_file_name=result_dict['originalFilename']
					outfile=os.path.join(output_directory,output_file_name)
					print("downloading %s" % result_dict.get('title'))
					resp, content = self.DRIVE._http.request(download_url)
					if resp.status == 200:
						if os.path.isfile(outfile):
							print("ERROR, %s already exist" % outfile)
						else:
							with open(outfile, 'wb') as f:
								f.write(content)
							print("OK")
					
			else:
				print('No Matches Found !!')
				exit()

		try:
			download_url=result_dict['downloadUrl']
		except Exception as e:
			print("No Download URL found for %s %s"%(file_name,file_id))
			exit()
		if download_url:
			output_file_name=result_dict['originalFilename']
			outfile=os.path.join(output_directory,output_file_name)
			print("downloading %s" % result_dict.get('title'))
			resp, content = self.DRIVE._http.request(download_url)
			if resp.status == 200:
				if os.path.isfile(outfile):
					print("ERROR, %s already exist" % outfile)
				else:
					with open(outfile, 'wb') as f:
						f.write(content)
					print("OK")
	def create_folder(self,folder_name,parent_id=''):
		#result_dict=self.retrieve_all_files(search_key=folder_name)
		items=self.search_file(search_key=folder_name,mimeType=FOLDER_MIME)
		if len(items):
			item=items[0]
			if parent_id:
				temp_dict=item['parents'][0]
				#print temp_dict
				if parent_id !=temp_dict['id']:
					print('Folder Already Excists (with matching parent ID)!!')
					return item
			# print('Folder Already Exsist!!')
			return item
		body = {'name': folder_name, 'mimeType': FOLDER_MIME,'title':folder_name}#, 'parents': [td_id]
		if parent_id:
			body['parents']=[parent_id]
		res= self.DRIVE.files().insert(body=body).execute()
		if res:
			# print('Folder created !!!with ID:',res['id'])
			return res
		return ''
	def delete_file(self,file_id='',file_name='',mimeType=''):
		if not file_id and file_name:
			items=self.search_file(search_key=file_name,mimeType=mimeType)
			print('len(items):',len(items))
			len_items=len(items)
			if len_items==1:
				item=items[0]
				file_id=item['id']
				result_dict=self.DRIVE.files().get(fileId=file_id).execute()
				file_name=result_dict['originalFilename']
				self.DRIVE.files().delete(fileId=item.get('id')).execute()
				print("%s deleted  Successfully!!!" % file_name)
				return True
			elif len_items>1:
				print('No of Matches found for \"'+file_name+'\" Count :'+str(len_items))
				for index,item in enumerate(items):
					print('Found file: %d : %s (%s) - %s' % (index,item.get('title'), item.get('id'),item.get('mimeType')))
				option=input('To delete all files Press \"Yes\" or Press the index number to delete : ')
				if option.lower()=='yes':
					for i in items:
						result_dict=self.DRIVE.files().get(fileId=file_id).execute()
						file_name=result_dict['originalFilename']
						self.DRIVE.files().delete(fileId=item.get('id')).execute()
						print("%s deleted  Successfully!!!" % file_name)
				elif option.isdigit():
					item=items[int(option)]
					print("item:",item)
					result_dict=self.DRIVE.files().get(fileId=file_id).execute()
					file_name=result_dict['originalFilename']
					self.DRIVE.files().delete(fileId=item.get('id')).execute()
					print("%s deleted  Successfully!!!" % file_name)
			else:
				print('No Matches Found !!')
				exit()
		elif file_id:
			result_dict=self.DRIVE.files().get(fileId=file_id).execute()
			file_name=result_dict['originalFilename']
		self.DRIVE.files().delete(fileId=file_id).execute()
		print("%s deleted  Successfully!!!" % file_name)
	def search_file(self,search_key='',mimeType='',is_folder=False):
		result=[]
		page_token = None
		q_string=''
		if is_folder:
			mimeType=FOLDER_MIME
		if search_key and mimeType:
			q_string="title=\'"+search_key+"\' and mimeType=\'"+mimeType+"\'"
		elif search_key and not mimeType:
			q_string="title=\'"+search_key+"\'"
		elif not search_key and  mimeType:
			q_string="mimeType=\'"+mimeType+"\'"
		file_count=0
		while True:
			if q_string:
				response = self.DRIVE.files().list(q=q_string,spaces='drive',pageToken=page_token).execute()
			else:
				response = self.DRIVE.files().list(spaces='drive',pageToken=page_token).execute()
			# print("len(response):",len(response))
			result.extend(response.get('items', []))
			for file in response.get('items', []):
				file_count+=1
				#print 'Found file: %d , %s (%s) - %s' % (file_count,file.get('title'), file.get('id'),file.get('mimeType'))
			page_token = response.get('nextPageToken', None)
			if page_token is None:
				break
		return result
	def convert_and_export(self,file_id='',file_name=''):
		data = self.DRIVE.files().export(fileId=file_id,mimeType='application/pdf').execute()
		#print 'data',data
		output_file_name = o_file_name+extension
		outfile=os.path.join(output_directory,output_file_name)
		print('outfile:',outfile)
		if os.path.isfile(outfile):
				print("ERROR, %s already exist" % outfile)
		else:
			with open(outfile, 'wb') as f:
				f.write(content)
			print("OK")
def process_image_to_text(file_lines,file_name ='',developer_mode =False):
	headers=['FILE_NAME','TRAN_ID','TRAN_DATE','TRAN_STATUS','BILLER_NAME','BILLER_ID','K_NUMBER','PAYMENT_MODE','AMOUNT','BILL_DATE','CONV_FEES','TOTAL_AMOUNT']

	tran_id =''
	tran_date =''
	tran_time =''
	tran_status =''

	biller_name =''
	biller_id =''
	k_number =''
	payment_mode =''
	amount =''
	bill_date =''
	conv_fees =''
	total_amount =''

	for index,each_line in enumerate(file_lines):
		if developer_mode:
			print(index,each_line)
			each_line
		
		t_date_pattern ='Txn. Date/Time (\w.+)'

		if not tran_id:
			pattern ='Txn. ID (\w+)\s(\d{8})'
			total_ids = re.findall(pattern, str(each_line),re.IGNORECASE)
			if total_ids :
				if developer_mode:
					print('tran_id total_ids :',total_ids)
				tran_id =' '.join([str(total_ids[0][0]),(total_ids[0][1])])
			if developer_mode:
				print('extracted tran_id :',tran_id)	
		if not tran_date:
			pattern ='Txn. Date/Time (\w.+)'
			total_ids = re.findall(pattern, str(each_line),re.IGNORECASE)
			if total_ids :
				if developer_mode:
					print('tran_date total_ids :',total_ids)
				tran_date =str(total_ids[0])
			if developer_mode:
				print('extracted tran_date :',tran_date)	
		if not tran_status:
			pattern ='Status (\w+)'
			total_ids = re.findall(pattern, str(each_line),re.IGNORECASE)
			if total_ids :
				if developer_mode:
					print('Status total_ids :',total_ids)
				tran_status =str(total_ids[0])
			if developer_mode:
				print('extracted tran_status :',tran_status)	

		if not biller_id:
			pattern ='(\w.+)\s(\w+)\s(\d{12})'
			total_ids = re.findall(pattern, str(each_line),re.IGNORECASE)
			if total_ids :
				if developer_mode:
					print('Biller id  total_ids :',total_ids)
				biller_name2 = str(total_ids[0][0]).strip('\n')
				biller_id = str(total_ids[0][1]).strip()
				k_number = str(total_ids[0][2]).strip()
				payment_mode = str(file_lines[index+1]).strip()
				biller_name = str(file_lines[index-1]).strip()+str(biller_name2)
			if developer_mode:
				print('extracted tran_status :',tran_status)
				print('extracted biller_name :',biller_name)
				print('extracted biller_id :',biller_id)
				print('extracted k_number :',k_number)
				print('extracted payment_mode :',payment_mode)
		if not amount:
			pattern ='(\d.00)'
			pattern2 ='\d.+[.]00'
			total_ids = re.findall(pattern, str(each_line),re.IGNORECASE)
			if developer_mode:
					print('Amount total_ids :',total_ids)
			if total_ids:
				splits = each_line.split()
				for i in splits:
					if '.00' in i.lower():
						amount_suspect = i
						if developer_mode:
							print('amount_suspect:',amount_suspect)
						second_extract = re.findall(pattern2, str(amount_suspect),re.IGNORECASE)
						if developer_mode:	
							print('second_extract:',second_extract)	
						amount = second_extract[0]
						total_amount = amount
				if developer_mode:
					print('extracted amount :',amount)	
		if not bill_date:
			pattern ='\d{2}/\d{2}/\d{4}'
			total_ids = re.findall(pattern, str(each_line),re.IGNORECASE)
			if total_ids :
				if developer_mode:
					print('bill date total_ids:',total_ids)
				bill_date =str(total_ids[0])
			if developer_mode:
				print('extracted bill_date :',bill_date)	
		# if  'txn. id' in each_line.lower():
		# 	t_id = each_line.replace('Txn. ID ','').strip()

		# elif  'status' in each_line.lower():
		# 	t_id = each_line.replace('Status ','').strip()


	res ={
		'tran_id' :tran_id
		,'tran_date' :tran_date
		,'tran_time' :tran_time
		,'tran_status' :tran_status

		,'biller_name' :biller_name
		,'biller_id' :biller_id
		,'k_number' :k_number
		,'payment_mode' :payment_mode

		,'amount' :amount
		,'bill_date' :bill_date
		,'conv_fees' :conv_fees
		,'total_amount' :total_amount
		,'file_name' :file_name
		}

	# # print ('res :',res)
	# full_log = [res['file_name'],res['tran_id'], res['tran_date']#, res['tran_time']
	# 			,res['tran_status'],
	# 			res['biller_name'], res['biller_id'],res['k_number'],res['payment_mode'], res['amount'] ,
	# 			res['bill_date'], res['conv_fees'],res['total_amount']
	# 			]
	# full_log_text = '\t'.join(apply_to_list(full_log, make_string=True)) + '\n'
	# write_into_file(file_name='Google_pay_log.txt', contents=add_timestamp(full_log_text), mode='a')
	# result_list.append(res)

	full_log = [res['file_name'],res['tran_id'], res['tran_date'], #res['tran_time']
				res['tran_status'],
				res['biller_name'], res['biller_id'],res['k_number'],res['payment_mode'], res['amount'] ,
				res['bill_date'], res['conv_fees'],res['total_amount']
				]
	full_log_text = apply_to_list(full_log, make_string=True)
	# if developer_mode:
	
	
	dt_string = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
	# print("date and time =", dt_string)
	full_log_text_log = full_log_text.copy()
	full_log_text_log.insert(0, str(dt_string))
	headers_log = headers.copy()
	headers_log.insert(0,'TIMESTAMP')

	# full_log_text_log = apply_to_list(full_log_text_log, make_string=True)
	print ('before writing full_log_text:',full_log_text)
	print ('Log text:',full_log_text_log)
	print ('Log text:',headers_log)
	
	write_excel(work_book_name='TOTAL_FILES_LOG.xlsx', rows=[full_log_text_log],headers = headers_log)
	write_excel(work_book_name='Extracted_text.xlsx', rows=[full_log_text],headers = headers)
	return res

def google_pay_format_data(input_directory,input_files = [],developer_mode = False):
	if developer_mode:
		print ('google_pay_format_data called!')
	result_list = []
	file_count =0
	if not input_files :
		for root, dirs, files in os.walk(input_directory):
			input_files = []
			for file_name in files:
				input_files.append(file_name)
			break

	try:
		for each_file_name in input_files:
			file_count+=1
			print ('Processing Extraction loop  :{0}/{1},  {2}'.format(file_count,len(input_files),each_file_name))
			
			each_file=os.path.join(input_directory,each_file_name)
			# file_lines = get_file_content(each_file,return_lines=True)
			
			file_lines = open(each_file).readlines()
			# if developer_mode:
			print ('File lines :',len(file_lines))
			result_list.append(process_image_to_text(file_lines,file_name = each_file_name))
		text_status = True
	except Exception as e:
		print('Error while extracting data from the text file :',e)
		text_status = False
	
	print (' Text Conversion: ',text_status)

	return {'results':result_list}

def apply_to_list(input_list,make_upper = False, make_lower= False,make_int= False,make_string=False):
    # added to make the list comprehension process easy  , added on 15/10/2020
    temp_list = input_list
    if make_string: return [str(i) for i in temp_list]
    elif make_upper: return [str(i).upper() for i in temp_list]
    elif make_lower: return [str(i).lower() for i in temp_list]
    elif make_int: return [int(i) for i in temp_list]
    return temp_list



if __name__ =='__main__':
	# upload_user_input = input('Do you want to upload/download images ( yes/no)')
	# text_extraction_user_input = input('Do you want to extract text ( yes/no)')
	# directory = input('Input directory :')


	# upload_user_input = 'yes'
	# text_extraction_user_input = 'yes'
	directory = sys.argv[1]
	

	client_key = 'client_secret.json'
	if not os.path.exists(client_key):
		print('Copy the client json to the directory :',os.getcwd(),' and rename it to client_secret.json')
	
	gd=GoogleDrive(client_key_json=client_key)
	gd.upload_images(directory=directory,output_directory='',developer_mode = False)

	print('Check results files :','Extracted_text.xlsx')
	print('Check log files :','TOTAL_FILES_LOG.xlsx')
	# if str(img_user_input).lower().strip == 'yes':
		
	# if str(text_extraction_user_input).lower().strip == 'yes':
	# 	full_results = google_pay_format_data(input_directory = text_file_directory,developer_mode = developer_mode)['results']
	# 	for res in full_results:
	# 		full_log = [res['file_name'],res['tran_id'], res['tran_date'], #res['tran_time']
	# 					res['tran_status'],
	# 					res['biller_name'], res['biller_id'],res['k_number'],res['payment_mode'], res['amount'] ,
	# 					res['bill_date'], res['conv_fees'],res['total_amount']
	# 					]
	# 		full_log_text = apply_to_list(full_log, make_string=True)
			# write_excel(work_book_name='converted_files.xlsx', rows=[full_log_text]):
	